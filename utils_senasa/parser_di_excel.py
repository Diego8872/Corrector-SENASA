"""
Parser de la Declaración de Importación (DI) en Excel.

Es la fuente PRINCIPAL de datos de la DI (más confiable que el PDF).
El Excel exportado por el sistema trae varias hojas; usamos:
  - "Carátula": datos generales del despacho (Aduana, Depósito, etc.)
  - "Item": datos por ítem (Origen, Procedencia, etc.)

Para la v1 tomamos el primer ítem con datos (ORIGEN no vacío), asumiendo
que Origen/Procedencia no varían entre ítems de un mismo despacho para
este chequeo. Si en el futuro hace falta comparar ítem por ítem, este
parser es el punto de extensión.
"""

import re
import openpyxl


def _strip_code(value):
    """'426 - REINO UNIDO' -> 'REINO UNIDO' ; '073 - EZEIZA' -> 'EZEIZA'"""
    if value is None:
        return None
    text = str(value).strip()
    m = re.match(r"^\s*\d+\s*-\s*(.+)$", text)
    return m.group(1).strip() if m else text


def extract_di_excel(xlsx_path: str) -> dict:
    """
    Devuelve un dict con: aduana, pais_origen, pais_procedencia, deposito.
    Cualquier campo no encontrado queda en None.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    data = {"aduana": None, "pais_origen": None, "pais_procedencia": None, "deposito": None}

    # --- Carátula ---
    if "Carátula" in wb.sheetnames:
        ws = wb["Carátula"]
        headers = [c.value for c in ws[1]]
        values = [c.value for c in ws[2]]
        header_map = {h: i for i, h in enumerate(headers) if h}

        if "ADUANA" in header_map:
            data["aduana"] = _strip_code(values[header_map["ADUANA"]])

    # --- Bultos (el Depósito real está acá, no en Carátula) ---
    if "Bultos" in wb.sheetnames:
        ws = wb["Bultos"]
        headers = [c.value for c in ws[1]]
        header_map = {h: i for i, h in enumerate(headers) if h}
        deposito_idx = header_map.get("DEPOSITO")
        if deposito_idx is not None:
            row2 = [c.value for c in ws[2]]
            if deposito_idx < len(row2):
                data["deposito"] = _strip_code(row2[deposito_idx])

    # --- Item (primer ítem con datos) ---
    if "Item" in wb.sheetnames:
        ws = wb["Item"]
        headers = [c.value for c in ws[1]]
        header_map = {h: i for i, h in enumerate(headers) if h}

        origen_idx = header_map.get("ORIGEN")
        procedencia_idx = header_map.get("PROCEDENCIA")

        for row in ws.iter_rows(min_row=2, values_only=True):
            origen_val = row[origen_idx] if origen_idx is not None else None
            if origen_val:
                data["pais_origen"] = _strip_code(origen_val)
                if procedencia_idx is not None:
                    data["pais_procedencia"] = _strip_code(row[procedencia_idx])
                break

    return data
